from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from spec_validator.models.spec import DataType, FieldSpec, NullPolicy, SpecDocument
from spec_validator.parsers.base import BaseSpecParser

_SCHEMA_SHEET_NAMES = {"schema", "fields", "spec", "field definitions", "data dictionary"}
_FIELD_NAME_HEADERS = {"field", "name", "column", "field_name", "column_name", "fieldname"}
_TYPE_HEADERS = {"type", "data_type", "datatype", "dtype"}
_NULLABLE_HEADERS = {"nullable", "required", "mandatory", "null", "optional"}
_ALLOWED_HEADERS = {"allowed_values", "values", "enum", "valid_values", "allowed"}
_DESCRIPTION_HEADERS = {"description", "desc", "notes", "note", "comment"}
_PATTERN_HEADERS = {"pattern", "regex", "format"}

_TYPE_MAP: dict[str, DataType] = {
    "string": DataType.STRING, "str": DataType.STRING, "text": DataType.STRING,
    "varchar": DataType.STRING, "char": DataType.STRING,
    "integer": DataType.INTEGER, "int": DataType.INTEGER, "long": DataType.INTEGER,
    "bigint": DataType.INTEGER, "smallint": DataType.INTEGER,
    "float": DataType.FLOAT, "double": DataType.FLOAT, "decimal": DataType.FLOAT,
    "numeric": DataType.FLOAT, "number": DataType.FLOAT,
    "boolean": DataType.BOOLEAN, "bool": DataType.BOOLEAN, "bit": DataType.BOOLEAN,
    "date": DataType.DATE,
    "datetime": DataType.DATETIME, "timestamp": DataType.DATETIME,
    "email": DataType.EMAIL,
    "url": DataType.URL, "uri": DataType.URL,
    "enum": DataType.ENUM,
}


class ExcelSpecParser(BaseSpecParser):
    def can_parse(self, path: str) -> bool:
        return Path(path).suffix.lower() in (".xlsx", ".xls")

    def parse(self, path: str, spec_id: str) -> SpecDocument:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet_name, rows = self._find_schema_sheet(wb)
        fields = self._rows_to_field_specs(rows)
        raw_content = self._serialize_to_text(wb)

        return SpecDocument(
            spec_id=spec_id,
            source_path=str(Path(path).resolve()),
            source_format="xlsx",
            title=Path(path).stem,
            fields=fields,
            raw_content=raw_content,
            parsed_at=datetime.now(timezone.utc).isoformat(),
        )

    def _find_schema_sheet(self, wb: openpyxl.Workbook) -> tuple[str, list[dict]]:
        for name in wb.sheetnames:
            if name.lower().strip() in _SCHEMA_SHEET_NAMES:
                rows = self._sheet_to_dicts(wb[name])
                if rows:
                    return name, rows

        # Fall back to first sheet with recognisable field-name column header
        for name in wb.sheetnames:
            rows = self._sheet_to_dicts(wb[name])
            if rows and any(
                k.lower().strip() in _FIELD_NAME_HEADERS for k in rows[0]
            ):
                return name, rows

        # Last resort: first sheet as-is
        first = wb.sheetnames[0]
        return first, self._sheet_to_dicts(wb[first])

    def _sheet_to_dicts(self, ws: Worksheet) -> list[dict]:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(c).strip() if c is not None else "" for c in rows[0]]
        result = []
        for row in rows[1:]:
            if all(c is None for c in row):
                continue
            result.append({
                headers[i]: (str(v).strip() if v is not None else "")
                for i, v in enumerate(row)
                if i < len(headers)
            })
        return result

    def _rows_to_field_specs(self, rows: list[dict]) -> list[FieldSpec]:
        if not rows:
            return []

        def _find_key(candidates: set[str]) -> str | None:
            for k in rows[0]:
                normalized = k.lower().strip().replace(" ", "_")
                if normalized in candidates:
                    return k
            return None

        name_key = _find_key(_FIELD_NAME_HEADERS)
        type_key = _find_key(_TYPE_HEADERS)
        nullable_key = _find_key(_NULLABLE_HEADERS)
        allowed_key = _find_key(_ALLOWED_HEADERS)
        desc_key = _find_key(_DESCRIPTION_HEADERS)
        pattern_key = _find_key(_PATTERN_HEADERS)

        if name_key is None:
            return []

        fields = []
        for row in rows:
            name = row.get(name_key, "").strip()
            if not name:
                continue

            raw_type = row.get(type_key, "").lower().strip() if type_key else ""
            data_type = _TYPE_MAP.get(raw_type, DataType.ANY)

            raw_nullable = row.get(nullable_key, "").lower().strip() if nullable_key else ""
            if raw_nullable in ("yes", "true", "1", "nullable", "null"):
                nullable = NullPolicy.NULLABLE
            elif raw_nullable in ("no", "false", "0", "required", "mandatory", "not null"):
                nullable = NullPolicy.REQUIRED
            else:
                nullable = NullPolicy.OPTIONAL

            raw_allowed = row.get(allowed_key, "").strip() if allowed_key else ""
            allowed_values = None
            if raw_allowed:
                allowed_values = [v.strip() for v in raw_allowed.split(",") if v.strip()]

            description = row.get(desc_key, "").strip() if desc_key else ""
            pattern = row.get(pattern_key, "").strip() if pattern_key else ""

            fields.append(FieldSpec(
                name=name,
                data_type=data_type,
                nullable=nullable,
                allowed_values=allowed_values or None,
                pattern=pattern or None,
                description=description,
            ))
        return fields

    def _serialize_to_text(self, wb: openpyxl.Workbook) -> str:
        parts = []
        for name in wb.sheetnames:
            ws = wb[name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            parts.append(f"## Sheet: {name}\n")
            for row in rows:
                cells = [str(c) if c is not None else "" for c in row]
                parts.append("| " + " | ".join(cells) + " |")
            parts.append("")
        return "\n".join(parts)
